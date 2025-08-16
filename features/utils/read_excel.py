from openpyxl import load_workbook

def read_login_data(excel_path, sheet_name):
    workbook = load_workbook(excel_path)
    sheet = workbook[sheet_name]

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
        data.append((row[0], row[1]))  # (userid, password)
    return data


def read_excel_as_dict_of_dicts(file_path, sheet_name, key_column="TCName"):
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]

    headers = [cell.value for cell in sheet[1]]
    key_index = headers.index(key_column)

    data = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        key = row[key_index]
        row_dict = {headers[i]: row[i] for i in range(len(headers)) if i != key_index}
        data[key] = row_dict

    return data


#login_data_dict  = read_excel_as_dict_of_dicts("../TestData/testdata.xlsx", "LoginData")
#print(login_data_dict)

#print(login_data_dict['test_verify_valid_login_TC04']['username'])